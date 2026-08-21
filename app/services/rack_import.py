from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (
    RACK_FACE_BACK,
    RACK_FACE_FRONT,
    RACK_MOUNT_RU,
    RACK_MOUNT_SIDE_PDU,
    Rack,
    RackItem,
    RackItemCategory,
    RackItemType,
    RackSite,
)

log = logging.getLogger("switcheroo.rack_import")

SEED_XLSX = Path(__file__).resolve().parent.parent / "seed_data" / "Brisbane_Albert_St_Rack_Layout.xlsx"

# Category name -> silhouette key
CATEGORY_DEFS: list[tuple[str, str, int]] = [
    ("Switch", "switch", 10),
    ("Patch panel", "patch", 20),
    ("Cable management", "cable", 30),
    ("Server", "server", 40),
    ("Mini PC", "mini", 50),
    ("Storage", "storage", 60),
    ("UPS", "ups", 70),
    ("PDU", "pdu", 80),
    ("Shelf", "shelf", 90),
    ("Blanking", "blank", 100),
    ("Telco / WAN", "telco", 110),
    ("Firewall / SD-WAN", "fw", 120),
    ("Console", "console", 130),
    ("Reserve", "reserve", 140),
    ("Other", "generic", 200),
]


def _classify(label: str) -> tuple[str, str, int, str, str]:
    """Return category, type_name, default_ru, face, mount."""
    text = " ".join(label.split())
    lower = text.lower()
    face = RACK_FACE_FRONT
    mount = RACK_MOUNT_RU
    if "rear mounted" in lower or "mounted at the back" in lower or "mounted at the rear" in lower:
        face = RACK_FACE_BACK
    if "front mounted" in lower:
        face = RACK_FACE_FRONT

    if "basic pdu" in lower or (lower.strip() == "pdu") or "vertical pdu" in lower:
        return "PDU", text, 0, RACK_FACE_BACK, RACK_MOUNT_SIDE_PDU
    if "blanking" in lower or lower.strip() == "blank":
        return "Blanking", "Blanking - Spare", 1, face, mount
    if "cable management" in lower:
        return "Cable management", "Cable Management", 1, face, mount
    if "shelf" in lower or "shelves" in lower:
        return "Shelf", "Shelves", 1, face, mount
    if "reserve" in lower:
        return "Reserve", text, 1, face, mount
    if "patch" in lower or "pp2" in lower or "rj45 patch" in lower:
        return "Patch panel", text, 1, face, mount
    if "cisco 9500" in lower or "cisco 9300" in lower or "catalyst" in lower:
        return "Switch", text, 1, face, mount
    if "dl360" in lower or "server" in lower:
        return "Server", text, 1, face, mount
    if "z2 mini" in lower:
        return "Mini PC", text, 1, face, mount
    if "netapp" in lower or "ds212" in lower or "fas " in lower or "filer" in lower:
        return "Storage", text, 1, face, mount
    if "ups" in lower:
        return "UPS", text, 3, face, mount
    if "opengear" in lower:
        return "Console", text, 1, RACK_FACE_BACK, mount
    if "telstra" in lower or "tpg" in lower or "ntu" in lower or "wan fibre" in lower or "wan fiber" in lower:
        return "Telco / WAN", text, 1, face, mount
    if "aruba" in lower or "asa" in lower or "riverbed" in lower or "sdwan" in lower or "sd-wan" in lower:
        return "Firewall / SD-WAN", text, 1, face, mount
    return "Other", text, 1, face, mount


def _ensure_categories(db: Session) -> dict[str, RackItemCategory]:
    out: dict[str, RackItemCategory] = {}
    for name, silhouette, order in CATEGORY_DEFS:
        cat = db.scalar(select(RackItemCategory).where(RackItemCategory.name == name))
        if cat is None:
            cat = RackItemCategory(name=name, silhouette=silhouette, sort_order=order)
            db.add(cat)
            db.flush()
        out[name] = cat
    return out


def _ensure_type(
    db: Session,
    categories: dict[str, RackItemCategory],
    category_name: str,
    type_name: str,
    *,
    default_ru: int,
    default_face: str,
    default_mount: str,
) -> RackItemType:
    cat = categories[category_name]
    row = db.scalar(
        select(RackItemType).where(RackItemType.category_id == cat.id, RackItemType.name == type_name)
    )
    if row is None:
        row = RackItemType(
            category_id=cat.id,
            name=type_name,
            default_ru_height=max(0, default_ru),
            default_face=default_face,
            default_mount=default_mount,
            default_network_ports=48 if category_name == "Switch" else (24 if category_name == "Patch panel" else 0),
            default_power_ports=2 if category_name in {"Server", "Storage", "UPS", "Switch"} else 0,
        )
        db.add(row)
        db.flush()
    return row


def _ru_number(cell: Any) -> int | None:
    if cell is None:
        return None
    text = str(cell).strip().upper()
    m = re.match(r"^RU\s*(\d+)$", text)
    if not m:
        return None
    return int(m.group(1))


def _is_filler(label: str) -> bool:
    """Repeated filler rows are one entry per RU in the sheet, not one tall device.

    Merging them would mean deleting a 27U blank just to place one server, so
    each RU stays independently replaceable.
    """
    lower = label.lower()
    return any(
        word in lower
        for word in ("blanking", "spare", "shelves", "shelf", "cable management", "reserve")
    )


def _merge_spans(cells: dict[int, str]) -> list[tuple[int, int, str]]:
    """cells: ru -> label. Return (ru_start_top, height, label). Doc: high RU at top."""
    if not cells:
        return []
    rus = sorted(cells.keys(), reverse=True)
    spans: list[tuple[int, int, str]] = []
    i = 0
    while i < len(rus):
        ru = rus[i]
        label = cells[ru]
        height = 1
        j = i + 1
        # identical labels continue downward (lower RU numbers)
        if not _is_filler(label):
            while j < len(rus) and rus[j] == ru - height and cells[rus[j]] == label:
                height += 1
                j += 1
        # NetApp shelf name + model on next RU down
        if (
            j < len(rus)
            and rus[j] == ru - height
            and "netapp disk shelf" in label.lower()
            and "ds212" in cells[rus[j]].lower()
        ):
            height += 1
            j += 1
        if (
            j < len(rus)
            and rus[j] == ru - height
            and "fas " in label.lower()
            and "filer" in cells[rus[j]].lower()
        ):
            height += 1
            j += 1
        spans.append((ru, height, label))
        i = j

    # Extend UPS into empty RUs below until next occupied or RU1
    occupied = set()
    for start, height, _ in spans:
        for r in range(start - height + 1, start + 1):
            occupied.add(r)
    extended: list[tuple[int, int, str]] = []
    for start, height, label in spans:
        if "ups" in label.lower():
            bottom = start - height + 1
            while bottom > 1 and (bottom - 1) not in occupied and (bottom - 1) not in cells:
                bottom -= 1
                height = start - bottom + 1
            for r in range(bottom, start + 1):
                occupied.add(r)
        extended.append((start, height, label))
    return extended


def _read_sheet_racks(ws, rack_defs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """rack_defs: name, label_col, item_col, optional side_pdu cols."""
    results = []
    for spec in rack_defs:
        cells: dict[int, str] = {}
        max_ru = 1
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            label_cell = row[spec["label_col"] - 1].value if spec["label_col"] <= len(row) else None
            ru = _ru_number(label_cell)
            if ru is None:
                continue
            max_ru = max(max_ru, ru)
            item_cell = row[spec["item_col"] - 1].value if spec["item_col"] <= len(row) else None
            if item_cell is None:
                continue
            text = str(item_cell).strip()
            if not text:
                continue
            cells[ru] = text
        side_items: list[dict[str, Any]] = []
        for side_spec in spec.get("side_pdus") or []:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                label_cell = row[spec["label_col"] - 1].value if spec["label_col"] <= len(row) else None
                ru = _ru_number(label_cell)
                if ru is None:
                    continue
                max_ru = max(max_ru, ru)
                val = row[side_spec["col"] - 1].value if side_spec["col"] <= len(row) else None
                if val is None or not str(val).strip():
                    continue
                side_items.append(
                    {
                        "ru": ru,
                        "label": str(val).strip(),
                        "side": side_spec["side"],
                    }
                )
        results.append(
            {
                "name": spec["name"],
                "floor": spec.get("floor") or "",
                "room": spec.get("room") or "",
                "sort_order": spec.get("sort_order") or 0,
                "ru_height": max(max_ru, spec.get("ru_height") or 45),
                "spans": _merge_spans(cells),
                "side_items": side_items,
            }
        )
    return results


def parse_brisbane_workbook(path: Path | None = None) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to import rack layouts") from exc

    xlsx = path or SEED_XLSX
    if not xlsx.is_file():
        raise FileNotFoundError(f"Rack layout workbook not found: {xlsx}")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    site = {
        "name": "Brisbane Albert St",
        "notes": "Imported from Brisbane_Albert_St_Rack_Layout (move-in layout; editable).",
    }
    rooms: list[dict[str, Any]] = []

    mcr = wb["Rack Layout MCR L27"]
    rooms.append(
        {
            "floor": "L27",
            "room": "MCR",
            "racks": _read_sheet_racks(
                mcr,
                [
                    {"name": "Server Rack B", "label_col": 1, "item_col": 2, "sort_order": 10, "floor": "L27", "room": "MCR"},
                    {"name": "Server Rack A", "label_col": 4, "item_col": 5, "sort_order": 20, "floor": "L27", "room": "MCR"},
                    {"name": "Comms Rack", "label_col": 7, "item_col": 8, "sort_order": 30, "floor": "L27", "room": "MCR"},
                    {
                        "name": "Floor distribution rack",
                        "label_col": 10,
                        "item_col": 11,
                        "sort_order": 40,
                        "floor": "L27",
                        "room": "MCR",
                    },
                ],
            ),
        }
    )

    l26 = wb["FDR Rack Layout L26"]
    rooms.append(
        {
            "floor": "L26",
            "room": "IDF",
            "racks": _read_sheet_racks(
                l26,
                [
                    {
                        "name": "FDR L26",
                        "label_col": 1,
                        "item_col": 3,
                        "sort_order": 10,
                        "floor": "L26",
                        "room": "IDF",
                        "side_pdus": [
                            {"col": 2, "side": "left"},
                            {"col": 4, "side": "right"},
                        ],
                    }
                ],
            ),
        }
    )

    l21 = wb["FDR Rack Layout L21"]
    rooms.append(
        {
            "floor": "L21",
            "room": "IDF",
            "racks": _read_sheet_racks(
                l21,
                [
                    {
                        "name": "FDR L21",
                        "label_col": 1,
                        "item_col": 3,
                        "sort_order": 10,
                        "floor": "L21",
                        "room": "IDF",
                        "side_pdus": [
                            {"col": 2, "side": "left"},
                            {"col": 4, "side": "right"},
                        ],
                    }
                ],
            ),
        }
    )
    return {"site": site, "rooms": rooms}


def import_brisbane_layout(db: Session, *, path: Path | None = None, force: bool = False) -> dict[str, int]:
    """Idempotent import. Skip if site already has racks unless force=True."""
    categories = _ensure_categories(db)
    parsed = parse_brisbane_workbook(path)
    site = db.scalar(select(RackSite).where(RackSite.name == parsed["site"]["name"]))
    created_racks = 0
    created_items = 0
    if site is None:
        site = RackSite(name=parsed["site"]["name"], notes=parsed["site"]["notes"], sort_order=10)
        db.add(site)
        db.flush()
    elif site.racks and not force:
        log.info("Rack site %s already has racks; skipping import", site.name)
        return {"sites": 0, "racks": 0, "items": 0, "types": 0}

    if force and site.racks:
        for rack in list(site.racks):
            db.delete(rack)
        db.flush()

    type_count_before = db.scalar(select(RackItemType.id).limit(1))

    for room in parsed["rooms"]:
        for rack_spec in room["racks"]:
            rack = db.scalar(
                select(Rack).where(Rack.site_id == site.id, Rack.name == rack_spec["name"])
            )
            if rack is None:
                rack = Rack(
                    site_id=site.id,
                    name=rack_spec["name"],
                    floor=rack_spec["floor"],
                    room=rack_spec["room"],
                    ru_height=rack_spec["ru_height"],
                    sort_order=rack_spec["sort_order"],
                )
                db.add(rack)
                db.flush()
                created_racks += 1

            existing = list(db.scalars(select(RackItem).where(RackItem.rack_id == rack.id)).all())
            if existing and not force:
                continue
            for item in existing:
                db.delete(item)
            db.flush()

            for ru_start, height, label in rack_spec["spans"]:
                cat_name, type_name, default_ru, face, mount = _classify(label)
                item_type = _ensure_type(
                    db,
                    categories,
                    cat_name,
                    type_name,
                    default_ru=max(height, default_ru),
                    default_face=face,
                    default_mount=mount,
                )
                use_height = height if height > 1 else max(height, item_type.default_ru_height or 1)
                if item_type.default_mount == RACK_MOUNT_SIDE_PDU:
                    use_height = 0
                db.add(
                    RackItem(
                        rack_id=rack.id,
                        item_type_id=item_type.id,
                        name=label,
                        ru_start=ru_start,
                        ru_height=max(1, use_height) if mount == RACK_MOUNT_RU else 0,
                        face=face,
                        mount=mount,
                        network_ports=item_type.default_network_ports,
                        power_ports=item_type.default_power_ports,
                    )
                )
                created_items += 1

            for side in rack_spec.get("side_items") or []:
                cat_name, type_name, default_ru, face, mount = _classify(side["label"])
                mount = RACK_MOUNT_SIDE_PDU
                item_type = _ensure_type(
                    db,
                    categories,
                    "PDU",
                    type_name if cat_name == "PDU" else side["label"],
                    default_ru=0,
                    default_face=RACK_FACE_BACK,
                    default_mount=mount,
                )
                db.add(
                    RackItem(
                        rack_id=rack.id,
                        item_type_id=item_type.id,
                        name=side["label"],
                        ru_start=side["ru"],
                        ru_height=0,
                        face=RACK_FACE_BACK,
                        mount=RACK_MOUNT_SIDE_PDU,
                        side=side["side"],
                        power_ports=24,
                    )
                )
                created_items += 1

    db.flush()
    types_now = len(list(db.scalars(select(RackItemType)).all()))
    log.info(
        "Rack import complete site=%s racks=%s items=%s types=%s",
        site.name,
        created_racks,
        created_items,
        types_now,
    )
    return {
        "sites": 1,
        "racks": created_racks,
        "items": created_items,
        "types": 0 if type_count_before else types_now,
    }
