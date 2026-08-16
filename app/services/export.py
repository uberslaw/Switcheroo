from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Port, Switch
from app.services.uptime import format_connected_for, short_if_name

HEADERS = (
    "Switch",
    "Port",
    "Purpose",
    "Label",
    "Status",
    "Admin",
    "VLAN",
    "VLAN name",
    "MAC",
    "IP",
    "ISE",
    "Connected uptime",
    "Last status poll",
    "Last detail poll",
)


def build_ports_workbook(db: Session, switches: list[Switch]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ports"
    header_font = Font(bold=True, color="1A1F12")
    header_fill = PatternFill("solid", fgColor="C6D36E")
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col, title)
        cell.font = header_font
        cell.fill = header_fill

    row_i = 2
    for switch in switches:
        ports = list(db.scalars(select(Port).where(Port.switch_id == switch.id).order_by(Port.if_index)).all())
        for port in ports:
            values = (
                switch.name,
                short_if_name(port.if_name),
                (port.purpose or "").replace("_", " "),
                port.friendly_label or "",
                port.oper_status,
                "shutdown" if port.admin_status == "down" else port.admin_status,
                port.vlan_id if port.vlan_id is not None else "",
                port.vlan_name or "",
                port.mac_address or "",
                port.ip_address or "",
                port.ise_status or "",
                format_connected_for(port.link_up_since),
                port.last_status_poll_at.strftime("%Y-%m-%d %H:%M:%S") if port.last_status_poll_at else "",
                port.last_detail_poll_at.strftime("%Y-%m-%d %H:%M:%S") if port.last_detail_poll_at else "",
            )
            for col, value in enumerate(values, start=1):
                ws.cell(row_i, col, value)
            row_i += 1

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 22
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(1, row_i - 1)}"
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
