from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select

from app.models import Switch
from app.services.export import HEADERS, build_ports_workbook
from tests.conftest import add_cs_user


def test_switch_page_renders_faceplate(client):
    client.post("/login", data={"username": "cs", "password": "cs"}, follow_redirects=False)
    home = client.get("/")
    assert "CS-BLD-A-AS01" in home.text
    page = client.get("/switches/1")
    assert page.status_code == 200
    assert "chassis" in page.text
    assert "Click a port on the switch" in page.text
    assert "Light green" in page.text
    assert "C9300-48" in page.text


def test_workbook_contains_expected_columns(seeded_db):
    switch = seeded_db.scalar(select(Switch).where(Switch.name == "CS-BLD-A-AS01"))
    payload = build_ports_workbook(seeded_db, [switch])
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert [cell.value for cell in ws[1]] == list(HEADERS)
    names = {ws.cell(row, 1).value for row in range(2, ws.max_row + 1)}
    assert names == {"CS-BLD-A-AS01"}
    assert ws.max_row == 49  # header + 48 ports
    assert any(ws.cell(row, 12).value and "Connected" in str(ws.cell(row, 12).value) for row in range(2, 10))


def test_cs_can_download_permitted_switch_xlsx(client, seeded_db):
    login = client.post("/login", data={"username": "cs", "password": "cs"}, follow_redirects=False)
    assert login.status_code == 303
    switch = seeded_db.scalar(select(Switch).where(Switch.name == "CS-BLD-A-AS01"))
    response = client.get(f"/switches/{switch.id}/export.xlsx")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert "CS-BLD-A-AS01" in response.headers["content-disposition"]
    wb = load_workbook(BytesIO(response.content))
    assert wb.active.max_row == 49


def test_cs_cannot_export_hidden_switch(client, seeded_db):
    add_cs_user(seeded_db, "cs-limited", "cs-limited", ["CS-BLD-A-AS01"])
    hidden = seeded_db.scalar(select(Switch).where(Switch.name == "CS-BLD-B-AS01"))
    client.post("/login", data={"username": "cs-limited", "password": "cs-limited"}, follow_redirects=False)
    response = client.get(f"/switches/{hidden.id}/export.xlsx", follow_redirects=False)
    assert response.status_code == 303
    assert response.headers["location"] == "/"


def test_visible_export_is_permission_aware(client, seeded_db):
    add_cs_user(seeded_db, "cs-limited", "cs-limited", ["CS-BLD-A-AS01"])
    client.post("/login", data={"username": "cs-limited", "password": "cs-limited"}, follow_redirects=False)
    response = client.get("/export.xlsx")
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    names = {wb.active.cell(row, 1).value for row in range(2, wb.active.max_row + 1)}
    assert names == {"CS-BLD-A-AS01"}
